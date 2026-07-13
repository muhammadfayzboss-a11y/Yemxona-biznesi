# -*- coding: utf-8 -*-
"""
Excel (xlsx) eksport. openpyxl kerak bo'ladi.
O'rnatish: pip install openpyxl
"""

import db
from formats import fmt_money


def export(path="qarzlar_hisoboti.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl o'rnatilmagan: pip install openpyxl")

    wb = Workbook()

    # --- 1-varaq: Qarzdorlar ---
    ws = wb.active
    ws.title = "Qarzdorlar"
    headers = ["#", "Mijoz", "Telefon", "Manzil", "Qolgan qarz (so'm)",
               "Qolgan qarz ($)", "Qarzdorga xabar"]
    ws.append(headers)
    debtors = db.debtors()
    for i, (c, u, d) in enumerate(debtors, 1):
        msg = (f"Assalomu alaykum, {c['name']} aka. Sizning qarzingiz: "
               f"{fmt_money(u, d)}. Batafsil: +998 __ ___ __ __")
        ws.append([i, c["name"], c["phone"] or "", c["address"] or "",
                   u, d, msg])

    # --- 2-varaq: Barcha operatsiyalar ---
    ws2 = wb.create_sheet("Operatsiyalar")
    h2 = ["Sana", "Mijoz", "Turi", "Summa (so'm)", "Summa ($)",
          "Mahsulot", "Muddat", "Izoh"]
    ws2.append(h2)
    type_names = {
        "sale": "Savdo",
        "payment": "To'lov",
        "return": "Qaytarish",
        "discount": "Chegirma",
    }
    for o in db.all_operations():
        ws2.append([
            o["created_at"], o["client_name"], type_names.get(o["type"], o["type"]),
            o["amount_uzs"] or 0, o["amount_usd"] or 0,
            o["product"] or "", o["due_date"] or "", o["note"] or "",
        ])

    # --- 3-varaq: Muddati o'tganlar ---
    ws3 = wb.create_sheet("Muddati o'tgan")
    ws3.append(["Mijoz", "Telefon", "Qolgan qarz", "Muddat"])
    for c, u, d, due in db.overdue_debtors():
        ws3.append([c["name"], c["phone"] or "", fmt_money(u, d), due])

    # Sarlavhalarni chiroyli qilish
    for sheet in (ws, ws2, ws3):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E7D32")
            cell.alignment = Alignment(horizontal="center")

    wb.save(path)
    return path


if __name__ == "__main__":
    print("Saqlandi:", export())
